"""
combined_script.py

This script combines the logic from the project's helper scripts:
1. Read tabular data from an Excel file and insert rows into the `lego_products` table.
2. Extract embedded images from the same Excel file and save them to
   `public/uploads/products/<id>/` (or `<name>/` depending on the id header).
3. Optionally write the first found image paths back into the DB (`--update-db`).
4. Export the `lego_products` table to `lego_products_export.json`.
5. Scan `public/uploads/products/` and update DB picture columns by matching folder
   names to product `name` (case-insensitive).

Credentials and configuration are read from environment variables. If a `.env`
file exists in the repository root and you have `python-dotenv` installed, it
will be loaded automatically so you can keep credentials there.

Usage examples:
  python backend/scripts/combined_script.py --xlsx "/workspaces/store/lego spreadsheet.xlsx" --id-header "Name" --update-db
  python backend/scripts/combined_script.py --help
"""

import os
import io
import uuid
import argparse
import json
import sys
from pathlib import Path

try:
	from dotenv import load_dotenv
	DOTENV_AVAILABLE = True
except Exception:
	DOTENV_AVAILABLE = False

import pandas as pd
try:
	from openpyxl import load_workbook
except Exception:
	raise SystemExit('Please install openpyxl: pip install openpyxl')

try:
	from PIL import Image
except Exception:
	raise SystemExit('Please install Pillow: pip install pillow')

import importlib.util
import os
spec = importlib.util.spec_from_file_location('supabase_rest', os.path.join(os.path.dirname(__file__), 'supabase_rest.py'))
supabase_rest = importlib.util.module_from_spec(spec)
spec.loader.exec_module(supabase_rest)

# --- repo paths ---
ROOT = Path(__file__).resolve().parents[2]
UPLOAD_BASE = ROOT / 'public' / 'uploads' / 'products'
UPLOAD_BASE.mkdir(parents=True, exist_ok=True)

def load_env():
	if DOTENV_AVAILABLE:
		env_path = ROOT / '.env'
		if env_path.exists():
			load_dotenv(env_path)

def get_db_config():
	password = os.getenv('PG_PASSWORD')
	if not password:
		raise RuntimeError('Database password must be set in environment variable PG_PASSWORD or .env file')
	return {
		'host': os.getenv('PG_HOST', 'localhost'),
		'port': os.getenv('PG_PORT', '5432'),
		'database': os.getenv('PG_DATABASE', 'lego_store'),
		'user': os.getenv('PG_USER', 'postgres'),
		'password': password,
	}

def create_table_if_not_exists(conn):
	# Schema creation is not performed via PostgREST here. Ensure your Supabase DB has the proper schema before running.
	print('Note: create_table_if_not_exists skipped (ensure DB schema exists in Supabase)')

def extract_data_from_excel(xlsx_path, id_header='id', skip_db_insert=False, dry_run=False):
	print(f'Reading spreadsheet: {xlsx_path}')
	df = pd.read_excel(xlsx_path)
	df.columns = [str(col).strip().replace(' ', '_').replace('\n', '_').replace('+', '_') for col in df.columns]
	id_header_clean = id_header.strip().replace(' ', '_')
	if id_header_clean not in df.columns:
		print(f"Warning: id header '{id_header}' not found in spreadsheet columns: {list(df.columns)}")
	if skip_db_insert:
		print('skip_db_insert set; skipping DB inserts')
		return len(df)
	# Insert/update rows via PostgREST upsert (on_conflict=id)
	rows = []
	for _, row in df.iterrows():
		id_val = None
		for key in (id_header_clean, 'id', 'ID', 'name', 'Name'):
			if key in row and pd.notna(row.get(key)):
				id_val = str(row.get(key))
				break
		if not id_val or id_val in ('nan', 'NaN'):
			id_val = str(uuid.uuid4())
		name = row.get('name') or row.get('Name') or None
		description = row.get('description') or row.get('Description') or None
		price = row.get('price_shipping_included') or row.get('price') or None
		if price is not None:
			try:
				price = float(str(price).replace('$','').replace(',','').strip())
			except Exception:
				price = None
		pieces = None
		if (row.get('lego_pieces') is not None) and (str(row.get('lego_pieces')).strip() != ''):
			try:
				pieces = int(row.get('lego_pieces'))
			except Exception:
				pieces = None
		rows.append({'id': id_val, 'name': name, 'description': description, 'price_shipping_included': price, 'lego_pieces': pieces, 'created_at': None, 'updated_at': None})
	# Batch upsert using on_conflict=id
	try:
		chunk = 100
		for i in range(0, len(rows), chunk):
			supabase_rest.upsert('lego_products', rows[i:i+chunk], params={'on_conflict': 'id'})
		print(f'Inserted/updated {len(rows)} rows via PostgREST')
	except Exception as e:
		print('Upsert failed:', e)
	return len(df)

def _image_bytes_from_openpyxl(img_obj):
	if hasattr(img_obj, '_data'):
		try:
			data = img_obj._data()
			if data:
				return data
		except Exception:
			pass
	if hasattr(img_obj, 'image') and img_obj.image is not None:
		bio = io.BytesIO()
		try:
			img_obj.image.save(bio, format='PNG')
			return bio.getvalue()
		except Exception:
			pass
	if hasattr(img_obj, 'path') and img_obj.path:
		try:
			with open(img_obj.path, 'rb') as f:
				return f.read()
		except Exception:
			pass
	raise RuntimeError('Could not extract image bytes from openpyxl image object')

def extract_images(xlsx_path, sheet_name=None, id_header='id', update_db=False, dry_run=False):
	print(f'Extracting embedded images from {xlsx_path} (sheet={sheet_name})')
	wb = load_workbook(xlsx_path, data_only=True)
	ws = wb[sheet_name] if sheet_name else wb.active
	headers = [cell.value for cell in ws[1]]
	headers = [h if h is None else str(h).strip().replace(' ', '_').replace('\n', '_') for h in headers]
	id_header_clean = id_header.strip().replace(' ', '_')
	if id_header_clean not in headers:
		print(f"Warning: id header '{id_header}' not found in worksheet headers: {headers}")
	id_col_idx = None
	try:
		id_col_idx = headers.index(id_header_clean) + 1
	except ValueError:
		id_col_idx = 1
	images = list(getattr(ws, '_images', []))
	if not images:
		print('No embedded images found in sheet')
		return {}
	mapping = {}
	for img in images:
		row_idx = None
		try:
			row_idx = int(img.anchor._from.row) + 1
		except Exception:
			try:
				anchor_from = getattr(img.anchor, 'from', None)
				if anchor_from and hasattr(anchor_from, 'row'):
					row_idx = int(anchor_from.row) + 1
				else:
					raise AttributeError
			except Exception:
				print('Could not determine image anchor row for an image, skipping')
				continue
		id_cell = ws.cell(row=row_idx, column=id_col_idx).value
		if id_cell is None:
			print(f'Row {row_idx} has no id cell, skipping image')
			continue
		product_id = str(id_cell)
		try:
			img_bytes = _image_bytes_from_openpyxl(img)
		except Exception as e:
			print(f'Failed to extract image bytes for row {row_idx}: {e}')
			continue
		folder = UPLOAD_BASE / product_id
		folder.mkdir(parents=True, exist_ok=True)
		fname = f"{uuid.uuid4().hex}.png"
		out_path = folder / fname
		if dry_run:
			print(f'[dry-run] would write image to {out_path}')
		else:
			with open(out_path, 'wb') as f:
				f.write(img_bytes)
		rel_url = f"/uploads/products/{product_id}/{fname}"
		mapping.setdefault(product_id, []).append(rel_url)
	if update_db and mapping and not dry_run:
		for pid, urls in mapping.items():
			rows = [{'product_id': pid, 'image_url': url} for url in urls]
			try:
				supabase_rest.insert('product_images', rows)
			except Exception as e:
				print('Failed to insert product_images for', pid, e)
	print('Image extraction complete. Products with images:', len(mapping))
	return mapping

def export_to_json(out_path='lego_products_export.json'):
	rows = supabase_rest.select('lego_products', params={'select': '*'})
	with open(out_path, 'w') as f:
		json.dump(rows, f, indent=2)
	print(f'Exported lego_products to {out_path} ({len(rows)} rows)')

def update_db_images_by_name(dry_run=False):
	if not UPLOAD_BASE.exists():
		print('No uploads directory, skipping update by folder name')
		return
	for folder in sorted(UPLOAD_BASE.iterdir()):
		if not folder.is_dir():
			continue
		product_name = folder.name
		exts = {'.png', '.jpg', '.jpeg', '.webp', '.gif'}
		files = [p for p in sorted(folder.iterdir()) if p.suffix.lower() in exts]
		if not files:
			continue
		urls = [f"/uploads/products/{folder.name}/{p.name}" for p in files]
		# Try exact match, then ilike
		rows = supabase_rest.select('lego_products', params={'select': 'id,name', 'name': f'eq.{product_name}'})
		if not rows:
			rows = supabase_rest.select('lego_products', params={'select': 'id,name', 'name': f'ilike.%{product_name}%', 'limit': '1'})
		if not rows:
			print(f'No product matched folder "{product_name}", skipping')
			continue
		product_id = rows[0]['id']
		for url in urls:
			if dry_run:
				print(f'[dry-run] Would add image for product id={product_id} name={rows[0].get("name")}: {url}')
			else:
				try:
					supabase_rest.insert('product_images', {'product_id': product_id, 'image_url': url})
				except Exception as e:
					print('Failed to insert product image', e)
		if not dry_run:
			print(f'Updated product id={product_id} name={rows[0].get("name")} with {len(urls)} images')

def main():
	load_env()
	parser = argparse.ArgumentParser(description='Combined import/export/image utility')
	parser.add_argument('--xlsx', required=True, help='Path to xlsx file')
	parser.add_argument('--sheet', help='Optional sheet name (defaults to active)')
	parser.add_argument('--id-header', default='id', help='Name of header column that holds product id (default: id)')
	parser.add_argument('--update-db', action='store_true', help='Update postgres lego_products table to reference saved images')
	parser.add_argument('--skip-db-insert', action='store_true', help='Skip inserting rows into DB (only extract images)')
	parser.add_argument('--dry-run', action='store_true', help='Do not write to DB or disk; just simulate')
	args = parser.parse_args()
	xlsx_path = Path(args.xlsx)
	if not xlsx_path.exists():
		print('Specified xlsx path does not exist:', xlsx_path)
		sys.exit(2)
	extract_data_from_excel(str(xlsx_path), id_header=args.id_header, skip_db_insert=args.skip_db_insert, dry_run=args.dry_run)
	mapping = extract_images(str(xlsx_path), sheet_name=args.sheet, id_header=args.id_header, update_db=args.update_db and not args.dry_run, dry_run=args.dry_run)
	if not args.dry_run:
		export_to_json()
	update_db_images_by_name(dry_run=args.dry_run)

if __name__ == '__main__':
	main()