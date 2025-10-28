"""
Extract embedded images from an XLSX file and save them to disk.

Usage:
  python extract_images_from_xlsx.py --xlsx path/to/file.xlsx [--sheet Sheet1] [--id-header id]

The script expects the first row of the sheet to be headers. It will locate the column whose header matches
`--id-header` (default: 'id') and use that cell in the same row as the image anchor to determine which product
the image belongs to.

It saves images to: public/uploads/products/<product_id>/ and prints a summary.
Optionally you can pass --update-db to write the first found image paths back into the `lego_products` table
(`pictures`, `pictures_1`, ... up to 5 slots). The script updates products via PostgREST (supabase REST helper);
no direct DB client is required.

Dependencies:
    pip install openpyxl pillow

"""
import os
import io
import uuid
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("Please install openpyxl: pip install openpyxl")

try:
    from PIL import Image
except ImportError:
    raise SystemExit("Please install Pillow: pip install pillow")

# psycopg2/direct DB access removed; this script updates products via PostgREST (supabase_rest helper)
import importlib.util
import os
spec = importlib.util.spec_from_file_location('supabase_rest', os.path.join(os.path.dirname(__file__), 'supabase_rest.py'))
supabase_rest = importlib.util.module_from_spec(spec)
spec.loader.exec_module(supabase_rest)

import pandas as pd
# psycopg2.sql was removed; using PostgREST via supabase_rest helper instead

# DB config (update or let environment variables override)
DB_CONFIG = {
    'host': os.environ.get('PG_HOST', 'localhost'),
    'port': os.environ.get('PG_PORT', '5432'),
    'database': os.environ.get('PG_DATABASE', 'lego_store'),
    'user': os.environ.get('PG_USER', 'postgres'),
    'password': os.environ.get('PG_PASSWORD', 'your_password'),
}

OUT_BASE = Path(__file__).resolve().parents[2] / 'public' / 'uploads' / 'products'
OUT_BASE.mkdir(parents=True, exist_ok=True)


def _get_image_bytes(img_obj):
    """Return image bytes (PNG) for an openpyxl Image object."""
    # openpyxl image objects may have ._data() that returns bytes OR .image (PIL Image)
    data = None
    if hasattr(img_obj, '_data'):
        try:
            data = img_obj._data()
        except Exception:
            data = None
    if data:
        return data

    if hasattr(img_obj, 'image') and img_obj.image is not None:
        pil = img_obj.image
        bio = io.BytesIO()
        pil.save(bio, format='PNG')
        return bio.getvalue()

    # Fallback: try to access .ref or .path (unlikely for embedded)
    if hasattr(img_obj, 'path') and img_obj.path:
        with open(img_obj.path, 'rb') as f:
            return f.read()

    raise RuntimeError('Could not extract image bytes from openpyxl image object')


def extract_images(xlsx_path, sheet_name=None, id_header='id', update_db=False):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Find header row (assume first row)
    headers = [cell.value for cell in ws[1]]
    if id_header not in headers:
        raise SystemExit(f"Header '{id_header}' not found in sheet headers: {headers}")
    id_col_idx = headers.index(id_header) + 1  # openpyxl is 1-indexed for columns

    images = list(getattr(ws, '_images', []))
    if not images:
        print('No embedded images found in sheet')
        return

    mapping = {}  # product_id -> list of saved paths

    for img in images:
        # Determine anchor row. Different openpyxl versions expose anchor differently.
        row = None
        try:
            # Many anchors expose .anchor._from.row (0-based)
            row = getattr(img.anchor, '_from').row + 1
        except Exception:
            # Try parsing anchor as a string like 'A5'
            try:
                coord = img.anchor._from
                row = coord.row + 1
            except Exception:
                # Last resort: try img.anchor._from.row
                try:
                    row = img.anchor.row
                except Exception:
                    print('Could not determine anchor row for an image; skipping')
                    continue

        # Read product id from the id column at that row
        product_id = ws.cell(row=row, column=id_col_idx).value
        if not product_id:
            print(f'Row {row} has no value in id column; skipping image')
            continue

        # Build output dir
        dest_dir = OUT_BASE / str(product_id)
        dest_dir.mkdir(parents=True, exist_ok=True)

        # Extract image bytes
        try:
            img_bytes = _get_image_bytes(img)
        except Exception as e:
            print('Failed to get image bytes for row', row, 'error:', e)
            continue

        # Save as PNG with uuid
        filename = f"{uuid.uuid4().hex}.png"
        out_path = dest_dir / filename
        with open(out_path, 'wb') as f:
            f.write(img_bytes)

        rel_url = f"/uploads/products/{product_id}/{filename}"
        mapping.setdefault(product_id, []).append(rel_url)
        print(f'Saved image for product {product_id} -> {rel_url}')

    # Optionally update DB: write up to 5 images into pictures..pictures_4
    if update_db:
        # Use PostgREST to patch each product's picture fields
        for pid, urls in mapping.items():
            fields = {}
            for i in range(min(5, len(urls))):
                key = 'pictures' if i == 0 else f'pictures_{i}'
                fields[key] = urls[i]
            try:
                awaitable = supabase_rest.patch('lego_products', fields, { 'id': f'eq.{pid}' })
            except Exception as e:
                print(f'Failed to update product {pid}:', e)
            else:
                print(f'Updated DB product {pid} with {len(urls)} image(s)')

    print('\nSummary:')
    for pid, urls in mapping.items():
        print(pid, len(urls))


# === Step 1: Read Excel file ===
def extract_data_from_excel(xlsx_path):
    df = pd.read_excel(xlsx_path)

    # Clean up column names (remove spaces/newlines)
    df.columns = [col.strip().replace(" ", "_").replace("\n", "_") for col in df.columns]

    # Insert data via PostgREST - assume schema already exists in DB
    rows = []
    for _, row in df.iterrows():
        rows.append({
            'name': row.get('Name'),
            'pictures': row.get('pictures'),
            'pictures_1': row.get('pictures.1'),
            'pictures_2': row.get('pictures.2'),
            'pictures_3': row.get('pictures.3'),
            'pictures_4': row.get('pictures.4'),
            'description': row.get('description'),
            'price_shipping_included': row.get('price+shipping_included'),
            'lego_pieces': int(row.get('lego_pieces')) if not pd.isna(row.get('lego_pieces')) else None,
        })
    try:
        # Batch insert
        chunk = 100
        for i in range(0, len(rows), chunk):
            supabase_rest.insert('lego_products', rows[i:i+chunk])
        print(f'Inserted {len(rows)} rows into lego_products via PostgREST')
    except Exception as e:
        print('Insert failed:', e)


if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--xlsx', required=True, help='Path to xlsx file')
    p.add_argument('--sheet', help='Optional sheet name (defaults to active)')
    p.add_argument('--id-header', default='id', help='Name of header column that holds product id (default: id)')
    p.add_argument('--update-db', action='store_true', help='Update postgres lego_products table to reference saved images')
    args = p.parse_args()

    # Extract data from Excel
    extract_data_from_excel(args.xlsx)

    # Existing image extraction logic
    extract_images(args.xlsx, sheet_name=args.sheet, id_header=args.id_header, update_db=args.update_db)
